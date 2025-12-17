"""
Output Generator - Generates RFP responses in multiple formats.

Generates three output formats:
1. JSON - Structured response for sales team review
2. Excel - Product recommendations with spec match, pricing, test costs
3. PDF - Professional proposal with comparison tables and pricing breakdown
"""
from typing import Dict, Any, List
from datetime import datetime
from pathlib import Path
import json
import structlog

# Excel generation imports
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF generation imports  
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

logger = structlog.get_logger()


class OutputGenerator:
    """
    Generates RFP response outputs in multiple formats.
    
    Output Formats:
    1. JSON: Structured data for sales team
    2. Excel: Product table with specs and pricing
    3. PDF: Professional proposal document
    """
    
    def __init__(self, output_dir: str = "outputs"):
        """Initialize Output Generator.
        
        Args:
            output_dir: Directory to save output files
        """
        self.logger = logger.bind(component="OutputGenerator")
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_all_formats(
        self,
        rfp_response: Dict[str, Any],
        rfp_id: int
    ) -> Dict[str, str]:
        """Generate all three output formats.
        
        Args:
            rfp_response: Consolidated RFP response from Master Agent
            rfp_id: RFP identifier
            
        Returns:
            Dictionary with paths to generated files
        """
        self.logger.info(
            "Generating output files in all formats",
            rfp_id=rfp_id
        )
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"rfp_response_{rfp_id}_{timestamp}"
        
        output_paths = {}
        
        # 1. Generate JSON
        json_path = self._generate_json(rfp_response, base_filename)
        output_paths['json'] = str(json_path)
        self.logger.info("JSON output generated", path=json_path)
        
        # 2. Generate Excel
        try:
            excel_path = self._generate_excel(rfp_response, base_filename)
            output_paths['excel'] = str(excel_path)
            self.logger.info("Excel output generated", path=excel_path)
        except Exception as e:
            self.logger.error(f"Excel generation failed: {e}")
            output_paths['excel'] = None
        
        # 3. Generate PDF
        try:
            pdf_path = self._generate_pdf(rfp_response, base_filename)
            output_paths['pdf'] = str(pdf_path)
            self.logger.info("PDF output generated", path=pdf_path)
        except Exception as e:
            self.logger.error(f"PDF generation failed: {e}")
            output_paths['pdf'] = None
        
        return output_paths
    
    def _generate_json(
        self,
        rfp_response: Dict[str, Any],
        base_filename: str
    ) -> Path:
        """Generate structured JSON response.
        
        Format includes:
        - RFP summary
        - Product recommendations with SKUs
        - Spec match percentages
        - Pricing breakdown
        - Test costs
        """
        json_path = self.output_dir / f"{base_filename}.json"
        
        # Structure for sales team review
        output = {
            "rfp_id": rfp_response.get('rfp_id'),
            "rfp_title": rfp_response.get('rfp_title'),
            "organization": rfp_response.get('organization'),
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_products": len(rfp_response.get('recommended_products', [])),
                "total_material_cost": rfp_response.get('material_costs', {}).get('net', 0),
                "total_testing_cost": rfp_response.get('testing_costs', {}).get('total', 0),
                "grand_total": rfp_response.get('total_costs', {}).get('grand_total', 0),
                "confidence_score": rfp_response.get('confidence_score', 0)
            },
            "product_recommendations": rfp_response.get('recommended_products', []),
            "material_costs": rfp_response.get('material_costs', {}),
            "testing_costs": rfp_response.get('testing_costs', {}),
            "total_costs": rfp_response.get('total_costs', {}),
            "spec_match_summary": rfp_response.get('spec_match_summary', {}),
            "compliance_summary": rfp_response.get('compliance_summary', {}),
            "processing_info": {
                "processed_at": rfp_response.get('processed_at'),
                "processing_time_seconds": rfp_response.get('processing_time_seconds', 0)
            }
        }
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        return json_path
    
    def _generate_excel(
        self,
        rfp_response: Dict[str, Any],
        base_filename: str
    ) -> Path:
        """Generate Excel file with product recommendations and pricing.
        
        Sheets:
        1. Summary - Overview and totals
        2. Products - Detailed product recommendations with spec match %
        3. Testing - Test requirements and costs
        4. Pricing - Complete pricing breakdown
        """
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl or pandas not installed, skipping Excel generation")
            raise ImportError("openpyxl and pandas required for Excel generation")
        
        excel_path = self.output_dir / f"{base_filename}.xlsx"
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("Summary", 0)
        self._create_summary_sheet(ws_summary, rfp_response)
        
        # 2. Products Sheet
        ws_products = wb.create_sheet("Products", 1)
        self._create_products_sheet(ws_products, rfp_response)
        
        # 3. Testing Sheet
        ws_testing = wb.create_sheet("Testing", 2)
        self._create_testing_sheet(ws_testing, rfp_response)
        
        # 4. Pricing Sheet
        ws_pricing = wb.create_sheet("Pricing", 3)
        self._create_pricing_sheet(ws_pricing, rfp_response)
        
        wb.save(excel_path)
        return excel_path
    
    def _create_summary_sheet(self, ws, rfp_response):
        """Create summary sheet in Excel."""
        # Header
        ws['A1'] = "RFP RESPONSE SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        
        # RFP Details
        row = 3
        ws[f'A{row}'] = "RFP Title:"
        ws[f'B{row}'] = rfp_response.get('rfp_title', 'N/A')
        row += 1
        ws[f'A{row}'] = "Organization:"
        ws[f'B{row}'] = rfp_response.get('organization', 'N/A')
        row += 1
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 2
        
        # Totals
        ws[f'A{row}'] = "FINANCIAL SUMMARY"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        total_costs = rfp_response.get('total_costs', {})
        ws[f'A{row}'] = "Material Cost:"
        ws[f'B{row}'] = total_costs.get('material', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "Testing Cost:"
        ws[f'B{row}'] = total_costs.get('testing', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "Subtotal:"
        ws[f'B{row}'] = total_costs.get('subtotal', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "GST (18%):"
        ws[f'B{row}'] = total_costs.get('gst', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "GRAND TOTAL:"
        ws[f'B{row}'] = total_costs.get('grand_total', 0)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '#,##0.00'
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_products_sheet(self, ws, rfp_response):
        """Create products sheet with recommendations."""
        # Headers
        headers = [
            "Item No",
            "Item Name",
            "OEM SKU",
            "Manufacturer",
            "Model",
            "Spec Match %",
            "Quantity",
            "Unit Price",
            "Line Total"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        products = rfp_response.get('recommended_products', [])
        for row_idx, product in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=product.get('item_number', 'N/A'))
            ws.cell(row=row_idx, column=2, value=product.get('item_name', 'N/A'))
            ws.cell(row=row_idx, column=3, value=product.get('oem_sku', 'N/A'))
            ws.cell(row=row_idx, column=4, value=product.get('manufacturer', 'N/A'))
            ws.cell(row=row_idx, column=5, value=product.get('model_number', 'N/A'))
            
            spec_match = ws.cell(row=row_idx, column=6, value=product.get('spec_match_%', 0))
            spec_match.number_format = '0.0"%"'
            
            ws.cell(row=row_idx, column=7, value=product.get('quantity', 0))
            
            unit_price = ws.cell(row=row_idx, column=8, value=product.get('unit_price', 0))
            unit_price.number_format = '#,##0.00'
            
            line_total = ws.cell(row=row_idx, column=9, value=product.get('line_total', 0))
            line_total.number_format = '#,##0.00'
        
        # Auto-width columns
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_testing_sheet(self, ws, rfp_response):
        """Create testing costs sheet."""
        # Headers
        ws['A1'] = "Test Type"
        ws['B1'] = "Description"
        ws['C1'] = "Cost"
        
        for col in ['A1', 'B1', 'C1']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[col].font = Font(color="FFFFFF", bold=True)
        
        # Data
        testing = rfp_response.get('testing_costs', {})
        row = 2
        
        if testing.get('routine', 0) > 0:
            ws[f'A{row}'] = "Routine Tests"
            ws[f'B{row}'] = "Routine tests as per standards"
            ws[f'C{row}'] = testing.get('routine', 0)
            ws[f'C{row}'].number_format = '#,##0.00'
            row += 1
        
        if testing.get('type', 0) > 0:
            ws[f'A{row}'] = "Type Tests"
            ws[f'B{row}'] = "Type tests at NABL lab"
            ws[f'C{row}'] = testing.get('type', 0)
            ws[f'C{row}'].number_format = '#,##0.00'
            row += 1
        
        if testing.get('acceptance', 0) > 0:
            ws[f'A{row}'] = "Acceptance Tests"
            ws[f'B{row}'] = "Acceptance tests at buyer site"
            ws[f'C{row}'] = testing.get('acceptance', 0)
            ws[f'C{row}'].number_format = '#,##0.00'
            row += 1
        
        row += 1
        ws[f'A{row}'] = "TOTAL TESTING COST"
        ws[f'C{row}'] = testing.get('total', 0)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
    
    def _create_pricing_sheet(self, ws, rfp_response):
        """Create complete pricing breakdown sheet."""
        # Title
        ws['A1'] = "COMPLETE PRICING BREAKDOWN"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Material Costs
        ws[f'A{row}'] = "MATERIAL COSTS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        material = rfp_response.get('material_costs', {})
        ws[f'A{row}'] = "Subtotal:"
        ws[f'B{row}'] = material.get('subtotal', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "Discount:"
        ws[f'B{row}'] = material.get('discount', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "Net Material:"
        ws[f'B{row}'] = material.get('net', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2
        
        # Testing Costs
        ws[f'A{row}'] = "TESTING COSTS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        testing = rfp_response.get('testing_costs', {})
        ws[f'A{row}'] = "Total Testing:"
        ws[f'B{row}'] = testing.get('total', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2
        
        # Grand Total
        ws[f'A{row}'] = "TOTAL (before tax)"
        ws[f'B{row}'] = rfp_response.get('total_costs', {}).get('subtotal', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "GST (18%)"
        ws[f'B{row}'] = rfp_response.get('total_costs', {}).get('gst', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "GRAND TOTAL"
        ws[f'B{row}'] = rfp_response.get('total_costs', {}).get('grand_total', 0)
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _generate_pdf(
        self,
        rfp_response: Dict[str, Any],
        base_filename: str
    ) -> Path:
        """Generate PDF proposal document.
        
        Sections:
        1. Cover page with RFP summary
        2. Product recommendations with comparison table
        3. Pricing breakdown
        4. Delivery timeline
        5. Terms and conditions
        """
        if not PDF_AVAILABLE:
            self.logger.warning("reportlab not installed, skipping PDF generation")
            raise ImportError("reportlab required for PDF generation")
        
        pdf_path = self.output_dir / f"{base_filename}.pdf"
        
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # 1. Cover Page
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph("RFP RESPONSE PROPOSAL", title_style))
        story.append(Spacer(1, 0.5*inch))
        
        cover_data = [
            ["RFP Title:", rfp_response.get('rfp_title', 'N/A')],
            ["Organization:", rfp_response.get('organization', 'N/A')],
            ["Prepared By:", "OEM Sales Team"],
            ["Date:", datetime.now().strftime("%B %d, %Y")],
        ]
        
        cover_table = Table(cover_data, colWidths=[2*inch, 4*inch])
        cover_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#366092')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        story.append(cover_table)
        story.append(PageBreak())
        
        # 2. Product Recommendations
        story.append(Paragraph("PRODUCT RECOMMENDATIONS", styles['Heading1']))
        story.append(Spacer(1, 0.2*inch))
        
        products = rfp_response.get('recommended_products', [])
        if products:
            product_headers = ["Item", "OEM SKU", "Manufacturer", "Spec Match", "Qty", "Unit Price", "Total"]
            product_data = [product_headers]
            
            for product in products:
                product_data.append([
                    product.get('item_name', 'N/A')[:20],
                    product.get('oem_sku', 'N/A'),
                    product.get('manufacturer', 'N/A'),
                    f"{product.get('spec_match_%', 0):.1f}%",
                    str(product.get('quantity', 0)),
                    f"Rs.{product.get('unit_price', 0):,.0f}",
                    f"Rs.{product.get('line_total', 0):,.0f}"
                ])
            
            product_table = Table(product_data, colWidths=[1.2*inch, 1*inch, 1*inch, 0.8*inch, 0.5*inch, 0.9*inch, 1*inch])
            product_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(product_table)
        
        story.append(Spacer(1, 0.3*inch))
        
        # 3. Pricing Breakdown
        story.append(Paragraph("PRICING BREAKDOWN", styles['Heading1']))
        story.append(Spacer(1, 0.2*inch))
        
        total_costs = rfp_response.get('total_costs', {})
        pricing_data = [
            ["Item", "Amount"],
            ["Material Cost", f"Rs.{total_costs.get('material', 0):,.2f}"],
            ["Testing Cost", f"Rs.{total_costs.get('testing', 0):,.2f}"],
            ["Subtotal", f"Rs.{total_costs.get('subtotal', 0):,.2f}"],
            ["GST (18%)", f"Rs.{total_costs.get('gst', 0):,.2f}"],
            ["GRAND TOTAL", f"Rs.{total_costs.get('grand_total', 0):,.2f}"],
        ]
        
        pricing_table = Table(pricing_data, colWidths=[4*inch, 2*inch])
        pricing_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F0F8')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(pricing_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 4. Delivery Timeline
        story.append(Paragraph("DELIVERY TIMELINE", styles['Heading1']))
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(
            "Standard delivery within 4-6 weeks from order confirmation. "
            "Express delivery available on request with additional charges.",
            styles['Normal']
        ))
        
        # Build PDF
        doc.build(story)
        
        return pdf_path
